from xhtml2pdf import pisa
from io import BytesIO
import openpyxl

def generate_pdf(html_content: str) -> bytes:
    """
    Generates a PDF from an HTML string using xhtml2pdf.
    """
    result = BytesIO()
    pisa_status = pisa.CreatePDF(
        html_content,
        dest=result
    )
    if pisa_status.err:
        raise Exception(f"Failed to generate PDF: {pisa_status.err}")
    
    return result.getvalue()

def generate_excel(data: list, headers: list = None) -> bytes:
    """
    Generates an Excel (.xlsx) file from a list of dicts.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    if not data:
        wb.save(result := BytesIO())
        return result.getvalue()

    if headers is None:
        headers = list(data[0].keys())

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=str(header))

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            # Convert dicts/lists to strings for Excel
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    result = BytesIO()
    wb.save(result)
    return result.getvalue()
